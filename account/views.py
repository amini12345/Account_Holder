from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count
from django.views.decorators.http import require_POST
from django.utils import timezone
from holder.models import Items, ItemHistory, PersonalInfo, ItemChangeRequest
from .forms import ItemForm
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from datetime import datetime
import json

@login_required
def dashboard(request):
    """صفحه داشبورد با آمار و نمودارهای کالاها"""
    # آمار کلی کالاها
    total_items = Items.objects.count()
    technical_items = Items.objects.filter(type_Item='Technical').count()
    non_technical_items = Items.objects.filter(type_Item='Non-technical').count()
    
    # آمار درخواست‌های تغییر
    total_change_requests = ItemChangeRequest.objects.count()
    pending_change_requests = ItemChangeRequest.objects.filter(status='pending').count()
    recent_change_requests = ItemChangeRequest.objects.filter(
        created_at__gte=timezone.now() - timezone.timedelta(days=7)
    ).order_by('-created_at')[:5]
    
    # آمار وضعیت کالاها - تبدیل به لیست برای استفاده در تمپلیت
    status_stats = []
    for status_code, status_name in Items.STATUS_CHOICES:
        count = Items.objects.filter(status_item=status_code).count()
        percentage = (count / total_items * 100) if total_items > 0 else 0
        
        # محاسبه زیر وضعیت‌ها برای این وضعیت
        sub_statuses = []
        if status_code in Items.STATUS_SUB_MAPPING:
            # محاسبه تعداد کل آیتم‌هایی که دارای زیر وضعیت هستند
            items_with_sub_status = Items.objects.filter(
                status_item=status_code, 
                status_sub_item__isnull=False
            ).count()
            
            # محاسبه تعداد آیتم‌هایی که زیر وضعیت ندارند
            items_without_sub_status = Items.objects.filter(
                status_item=status_code, 
                status_sub_item__isnull=True
            ).count()
            
            for sub_code, sub_name in Items.STATUS_SUB_MAPPING[status_code]:
                sub_count = Items.objects.filter(status_item=status_code, status_sub_item=sub_code).count()
                # درصد نسبت به کل کالاها
                percentage_total = (sub_count / total_items * 100) if total_items > 0 else 0
                # درصد نسبت به وضعیت بالادستی (فقط آیتم‌هایی که دارای زیر وضعیت هستند)
                percentage_parent = (sub_count / count * 100) if count > 0 else 0
                # درصد نسبت به آیتم‌هایی که دارای زیر وضعیت هستند
                percentage_parent_with_sub = (sub_count / items_with_sub_status * 100) if items_with_sub_status > 0 else 0
                
                sub_statuses.append({
                    'code': sub_code,
                    'name': sub_name,
                    'count': sub_count,
                    'percentage_total': round(percentage_total, 2),
                    'percentage_parent': round(percentage_parent, 2),
                    'percentage_parent_with_sub': round(percentage_parent_with_sub, 2)
                })
            
            # اضافه کردن آیتم‌هایی که زیر وضعیت ندارند
            if items_without_sub_status > 0:
                percentage_total_no_sub = (items_without_sub_status / total_items * 100) if total_items > 0 else 0
                percentage_parent_no_sub = (items_without_sub_status / count * 100) if count > 0 else 0
                
                sub_statuses.append({
                    'code': 'no_sub_status',
                    'name': 'بدون زیر وضعیت',
                    'count': items_without_sub_status,
                    'percentage_total': round(percentage_total_no_sub, 2),
                    'percentage_parent': round(percentage_parent_no_sub, 2),
                    'percentage_parent_with_sub': 0
                })
        
        status_stats.append({
            'code': status_code,
            'name': status_name,
            'count': count,
            'percentage': round(percentage, 2),
            'sub_statuses': sub_statuses
        })
    
    # دریافت اقلام کاربر جاری
    try:
        current_user = PersonalInfo.objects.get(Personnel_number=request.user.username)
        user_items = Items.objects.filter(PersonalInfo=current_user)
    except PersonalInfo.DoesNotExist:
        user_items = Items.objects.none()
        current_user = None
    
    # محاسبه مجموع برای اعتبارسنجی
    total_status_items = sum([status['count'] for status in status_stats])
    total_sub_status_items = 0
    for status in status_stats:
        for sub_status in status['sub_statuses']:
            total_sub_status_items += sub_status['count']
    
    # آماده‌سازی داده‌ها برای JavaScript
    dashboard_data = {
        'totalItems': total_items,
        'technicalItems': technical_items,
        'nonTechnicalItems': non_technical_items,
        'hasData': total_items > 0,
        'statusData': status_stats,
        'statusLabels': [status['name'] for status in status_stats],
        'statusCounts': [status['count'] for status in status_stats]
    }
    
    context = {
        'total_items': total_items,
        'technical_items': technical_items,
        'non_technical_items': non_technical_items,
        'status_stats': status_stats,
        'items': user_items,
        'user': current_user if current_user else request.user,
        'total_status_items': total_status_items,
        'total_sub_status_items': total_sub_status_items,
        'validation_check': total_items == total_status_items,
        'dashboard_data': dashboard_data,
        'total_change_requests': total_change_requests,
        'pending_change_requests': pending_change_requests,
        'recent_change_requests': recent_change_requests,
    }
    
    return render(request, 'registration/dashboard.html', context)

@login_required
def home(request):
    return render(request, 'registration/home.html')

class HomeView(LoginRequiredMixin, ListView):
    model = Items
    template_name = 'registration/home.html'
    context_object_name = 'object_list'
    paginate_by = 10
    
    def get_queryset(self):
        queryset = Items.objects.select_related('PersonalInfo').all()
        search_query = self.request.GET.get('search')
        brand_search = self.request.GET.get('brand_search')
        serial_search = self.request.GET.get('serial_search')
        code_search = self.request.GET.get('code_search')
        holder_search = self.request.GET.get('holder_search')
        type_filter = self.request.GET.get('type_filter')
        status_filter = self.request.GET.get('status_filter')
        sub_status_filter = self.request.GET.get('sub_status_filter')
        
        # جستجو در نام کالا
        if search_query:
            queryset = queryset.filter(Technical_items__icontains=search_query)
        
        # جستجو در برند
        if brand_search:
            queryset = queryset.filter(brand__icontains=brand_search)
        
        # جستجو در شماره سریال
        if serial_search:
            queryset = queryset.filter(serial_number__icontains=serial_search)
        
        # جستجو در کد محصول
        if code_search:
            queryset = queryset.filter(Product_code__icontains=code_search)
        
        # جستجو در دارنده حساب
        if holder_search:
            queryset = queryset.filter(
                Q(PersonalInfo__name__icontains=holder_search) |
                Q(PersonalInfo__family__icontains=holder_search) |
                Q(PersonalInfo__Personnel_number__icontains=holder_search)
            )
        
        # یلتر بر اساس نوع کالا
        if type_filter:
            queryset = queryset.filter(type_Item=type_filter)
        
        # فیلتر بر اساس وضعیت کالا
        if status_filter:
            queryset = queryset.filter(status_item=status_filter)
        
        # فیلتر بر اساس زیر وضعیت کالا
        if sub_status_filter:
            queryset = queryset.filter(status_sub_item=sub_status_filter)
        
        return queryset.order_by('-register_date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # آمار کالاها
        total_items = Items.objects.count()
        technical_items = Items.objects.filter(type_Item='Technical').count()
        non_technical_items = Items.objects.filter(type_Item='Non-technical').count()
        
        # آمار درخواست‌های تغییر
        pending_change_requests = ItemChangeRequest.objects.filter(status='pending').count()
        
        # لیست افراد برای انتقال دسته‌ای
        all_people = PersonalInfo.objects.all().order_by('name', 'family')
        
        context.update({
            'total_items': total_items,
            'technical_items': technical_items,
            'non_technical_items': non_technical_items,
            'pending_change_requests': pending_change_requests,
            'all_people': all_people,
            'search_query': self.request.GET.get('search', ''),
            'brand_search': self.request.GET.get('brand_search', ''),
            'serial_search': self.request.GET.get('serial_search', ''),
            'code_search': self.request.GET.get('code_search', ''),
            'holder_search': self.request.GET.get('holder_search', ''),
        })
        
        return context

class ItemCreateView(LoginRequiredMixin, CreateView):
    model = Items
    form_class = ItemForm
    template_name = 'registration/item_form.html'
    success_url = reverse_lazy('account:home')
    
    def form_valid(self, form):
        messages.success(self.request, 'کالا با موفقیت اضافه شد.')
        return super().form_valid(form)

class ItemUpdateView(LoginRequiredMixin, UpdateView):
    model = Items
    form_class = ItemForm
    template_name = 'registration/item_form.html'
    success_url = reverse_lazy('account:home')
    
    def form_valid(self, form):
        messages.success(self.request, 'کالا با موفقیت به‌روزرسانی شد.')
        return super().form_valid(form)

class ItemDeleteView(LoginRequiredMixin, DeleteView):
    model = Items
    template_name = 'registration/item_confirm_delete.html'
    success_url = reverse_lazy('account:home')
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'کالا با موفقیت حذف شد.')
        return super().delete(request, *args, **kwargs)

@login_required
def item_detail(request, pk):
    item = get_object_or_404(Items, pk=pk)
    history = ItemHistory.objects.filter(item=item).order_by('-action_date')
    
    context = {
        'item': item,
        'history': history,
    }
    
    return render(request, 'registration/item_detail.html', context)

@login_required
def ajax_search_items(request):
    """جستجوی AJAX برای کالاها"""
    search_query = request.GET.get('q', '')
    
    if search_query:
        items = Items.objects.filter(
            Q(Technical_items__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            Q(Product_code__icontains=search_query)
        )[:10]
        
        results = []
        for item in items:
            results.append({
                'id': item.id,
                'name': item.Technical_items or 'بدون نام',
                'serial': item.serial_number or 'ندارد',
                'code': item.Product_code or 'ندارد',
                'type': item.get_type_Item_display(),
            })
        
        return JsonResponse({'results': results})
    
    return JsonResponse({'results': []})

@login_required
def profiles_view(request):
    """نمایش پروفایل‌ها و کالاهای هر فرد"""
    selected_person_id = request.GET.get('person_id')
    search_query = request.GET.get('search', '')
    
    # دریافت افراد با امکان جستجو
    people_queryset = PersonalInfo.objects.all()
    
    if search_query:
        people_queryset = people_queryset.filter(
            Q(name__icontains=search_query) |
            Q(family__icontains=search_query) |
            Q(Personnel_number__icontains=search_query) |
            Q(National_ID__icontains=search_query)
        )
    
    people = people_queryset.order_by('name', 'family')
    
    # اگر فردی انتخاب شده، کالاهای او را دریافت کن
    selected_person = None
    person_items = None
    
    if selected_person_id:
        try:
            selected_person = PersonalInfo.objects.get(Personnel_number=selected_person_id)
            person_items = Items.objects.filter(PersonalInfo=selected_person).order_by('-register_date')
        except PersonalInfo.DoesNotExist:
            selected_person = None
            person_items = None
    
    # آمار افراد
    total_people = people_queryset.count()
    
    context = {
        'people': people,
        'selected_person': selected_person,
        'person_items': person_items,
        'selected_person_id': selected_person_id,
        'search_query': search_query,
        'total_people': total_people,
    }
    
    return render(request, 'registration/profiles.html', context)

@login_required
def history_view(request):
    """نمایش لیست کالاها برای انتخاب و مشاهده تاریخچه"""
    selected_item_id = request.GET.get('item_id')
    search_query = request.GET.get('search', '')
    type_filter = request.GET.get('type_filter', '')
    
    # دریافت کالاها با امکان جستجو و فیلتر
    items_queryset = Items.objects.select_related('PersonalInfo').all()
    
    if search_query:
        items_queryset = items_queryset.filter(
            Q(Technical_items__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            Q(Product_code__icontains=search_query) |
            Q(PersonalInfo__name__icontains=search_query) |
            Q(PersonalInfo__family__icontains=search_query)
        )
    
    # فیلتر بر اساس نوع کالا (فنی/غیر فنی)
    if type_filter:
        items_queryset = items_queryset.filter(type_Item=type_filter)
    
    items = items_queryset.order_by('-register_date')
    
    # اگر کالایی انتخاب شده، تاریخچه آن را دریافت کن
    selected_item = None
    item_history = None
    
    if selected_item_id:
        try:
            selected_item = Items.objects.get(id=selected_item_id)
            item_history = ItemHistory.objects.filter(item=selected_item).select_related(
                'from_person', 'to_person'
            ).order_by('-action_date')
        except Items.DoesNotExist:
            selected_item = None
            item_history = None
    
    # آمار کالاها برای نمایش در فیلتر
    total_items = items_queryset.count()
    technical_items = items_queryset.filter(type_Item='Technical').count()
    non_technical_items = items_queryset.filter(type_Item='Non-technical').count()
    
    context = {
        'items': items,
        'selected_item': selected_item,
        'item_history': item_history,
        'selected_item_id': selected_item_id,
        'search_query': search_query,
        'type_filter': type_filter,
        'total_items': total_items,
        'technical_items': technical_items,
        'non_technical_items': non_technical_items,
    }
    
    return render(request, 'registration/history.html', context)

@login_required
def export_excel_fields_selection(request):
    """صفحه انتخاب فیلدهای خروجی Excel"""
    context = {
        'search_query': request.GET.get('search', ''),
        'brand_search': request.GET.get('brand_search', ''),
        'serial_search': request.GET.get('serial_search', ''),
        'code_search': request.GET.get('code_search', ''),
        'holder_search': request.GET.get('holder_search', ''),
        'type_filter': request.GET.get('type_filter', ''),
        'status_filter': request.GET.get('status_filter', ''),
        'sub_status_filter': request.GET.get('sub_status_filter', ''),
    }
    return render(request, 'registration/excel_fields_selection.html', context)

@login_required
def generate_excel(request):
    """تولید فایل Excel با فیلدهای انتخاب شده"""
    if request.method != 'POST':
        return redirect('account:export_excel_fields_selection')
    
    selected_fields = request.POST.getlist('selected_fields')
    if not selected_fields:
        messages.error(request, 'لطفاً حداقل یک فیلد را انتخاب کنید.')
        return redirect('account:export_excel_fields_selection')
    
    # دریافت همان queryset که در HomeView استفاده می‌شود
    queryset = Items.objects.select_related('PersonalInfo').all()
    search_query = request.POST.get('search')
    brand_search = request.POST.get('brand_search')
    serial_search = request.POST.get('serial_search')
    code_search = request.POST.get('code_search')
    holder_search = request.POST.get('holder_search')
    type_filter = request.POST.get('type_filter')
    status_filter = request.POST.get('status_filter')
    sub_status_filter = request.POST.get('sub_status_filter')
    
    # اعمال فیلترها (همان منطق HomeView)
    if search_query:
        queryset = queryset.filter(Technical_items__icontains=search_query)
    if brand_search:
        queryset = queryset.filter(brand__icontains=brand_search)
    if serial_search:
        queryset = queryset.filter(serial_number__icontains=serial_search)
    if code_search:
        queryset = queryset.filter(Product_code__icontains=code_search)
    if holder_search:
        queryset = queryset.filter(
            Q(PersonalInfo__name__icontains=holder_search) |
            Q(PersonalInfo__family__icontains=holder_search) |
            Q(PersonalInfo__Personnel_number__icontains=holder_search)
        )
    if type_filter:
        queryset = queryset.filter(type_Item=type_filter)
    if status_filter:
        queryset = queryset.filter(status_item=status_filter)
    if sub_status_filter:
        queryset = queryset.filter(status_sub_item=sub_status_filter)
    
    queryset = queryset.order_by('-register_date')
    
    # تعریف تمام فیلدهای ممکن
    all_fields = {
        'row_number': {'header': 'ردیف', 'width': 8},
        'name': {'header': 'نام کالا', 'width': 25},
        'type': {'header': 'نوع کالا', 'width': 15},
        'brand': {'header': 'برند', 'width': 15},
        'configuration': {'header': 'پیکربندی', 'width': 30},
        'status': {'header': 'وضعیت کالا', 'width': 15},
        'sub_status': {'header': 'زیر وضعیت', 'width': 15},
        'serial': {'header': 'شماره سریال', 'width': 20},
        'product_code': {'header': 'کد محصول', 'width': 15},
        'holder': {'header': 'دارنده حساب', 'width': 25},
        'register_date': {'header': 'تاریخ ثبت', 'width': 20},
        'update_date': {'header': 'تاریخ بروزرسانی', 'width': 20},
    }
    
    # فیلتر کردن فیلدهای انتخاب شده
    selected_field_configs = [(field, all_fields[field]) for field in selected_fields if field in all_fields]
    
    # ایجاد workbook و worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "لیست کالاها"
    
    # تنظیم RTL (راست به چپ) برای worksheet
    ws.sheet_view.rightToLeft = True
    
    # تنظیم عرض ستون‌ها
    for i, (field_key, field_config) in enumerate(selected_field_configs, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = field_config['width']
    
    # استایل هدر
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # نوشتن هدرها
    for col, (field_key, field_config) in enumerate(selected_field_configs, 1):
        cell = ws.cell(row=1, column=col, value=field_config['header'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # نوشتن داده‌ها
    for row_num, item in enumerate(queryset, 2):
        for col, (field_key, field_config) in enumerate(selected_field_configs, 1):
            value = get_field_value(item, field_key, row_num - 1)
            ws.cell(row=row_num, column=col, value=value)
    
    # تنظیم alignment برای همه سلول‌ها
    for row in ws.iter_rows():
        for cell in row:
            if cell.row > 1:  # به جز هدر
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # تنظیم نام فایل
    current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename_parts = ['کالاها']
    
    # اضافه کردن فیلترها به نام فایل
    if search_query:
        filename_parts.append(f'نام_{search_query}')
    if brand_search:
        filename_parts.append(f'برند_{brand_search}')
    if serial_search:
        filename_parts.append(f'سریال_{serial_search}')
    if code_search:
        filename_parts.append(f'کد_{code_search}')
    if holder_search:
        filename_parts.append(f'دارنده_{holder_search}')
    if type_filter:
        type_name = 'فنی' if type_filter == 'Technical' else 'غیرفنی'
        filename_parts.append(f'نوع_{type_name}')
    if status_filter:
        status_names = {
            'hardware': 'سخت_افزار',
            'Delivery': 'تحویل',
            'warehouse': 'انبار'
        }
        filename_parts.append(f'وضعیت_{status_names.get(status_filter, status_filter)}')
    if sub_status_filter:
        sub_status_names = {
            'repair': 'تعمیر',
            'upgrade': 'ارتقا',
            'external': 'خارج',
            'internal': 'داخل',
            'ready': 'آماده_بکار',
            'returned_good': 'عودتی_سالم',
            'returned_worn': 'عودتی_فرسوده'
        }
        filename_parts.append(f'زیروضعیت_{sub_status_names.get(sub_status_filter, sub_status_filter)}')
    
    filename_parts.append(current_time)
    filename = '_'.join(filename_parts) + '.xlsx'
    
    # ایجاد response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # ذخیره workbook در response
    wb.save(response)
    
    return response

def get_field_value(item, field_key, row_index):
    """دریافت مقدار فیلد برای یک آیتم"""
    if field_key == 'row_number':
        return row_index + 1
    elif field_key == 'name':
        return item.Technical_items or 'تعریف نشده'
    elif field_key == 'type':
        return 'فنی' if item.type_Item == 'Technical' else 'غیر فنی' if item.type_Item == 'Non-technical' else 'نامشخص'
    elif field_key == 'brand':
        return item.brand or 'تعریف نشده'
    elif field_key == 'configuration':
        return item.Configuration or 'ندارد'
    elif field_key == 'status':
        status_display = {
            'hardware': 'سخت افزار (تعمیری)',
            'Delivery': 'تحویل',
            'warehouse': 'انبار',
            'Healthy': 'سالم',
            'Repairing': 'تعمیری', 
            'worn out': 'فرسوده',
            'other': 'سایر'
        }.get(item.status_item, 'نامشخص')
        return status_display
    elif field_key == 'sub_status':
        return item.get_status_sub_item_display() if item.status_sub_item else 'ندارد'
    elif field_key == 'serial':
        return item.serial_number or 'ندارد'
    elif field_key == 'product_code':
        return item.Product_code or 'ندارد'
    elif field_key == 'holder':
        if item.PersonalInfo:
            return f"{item.PersonalInfo.name} {item.PersonalInfo.family} ({item.PersonalInfo.Personnel_number})"
        else:
            return 'بدون دارنده'
    elif field_key == 'register_date':
        return str(item.jinfo)
    elif field_key == 'update_date':
        if item.update_date:
            return item.update_date.strftime('%Y/%m/%d %H:%M')
        else:
            return 'ندارد'
    else:
        return ''

@login_required
def export_excel(request):
    """خروجی Excel از لیست کالاها با در نظر گیری جستجو و فیلترها - نسخه قدیمی"""
    
    # دریا��ت همان queryset که در HomeView استفاده می‌شود
    queryset = Items.objects.select_related('PersonalInfo').all()
    search_query = request.GET.get('search')
    brand_search = request.GET.get('brand_search')
    serial_search = request.GET.get('serial_search')
    code_search = request.GET.get('code_search')
    holder_search = request.GET.get('holder_search')
    type_filter = request.GET.get('type_filter')
    status_filter = request.GET.get('status_filter')
    sub_status_filter = request.GET.get('sub_status_filter')
    
    # جستجو در نام کالا
    if search_query:
        queryset = queryset.filter(Technical_items__icontains=search_query)
    
    # جستجو در برند
    if brand_search:
        queryset = queryset.filter(brand__icontains=brand_search)
    
    # جستجو در شماره سریال
    if serial_search:
        queryset = queryset.filter(serial_number__icontains=serial_search)
    
    # جستجو در کد محصول
    if code_search:
        queryset = queryset.filter(Product_code__icontains=code_search)
    
    # ��ستجو در دارنده حساب
    if holder_search:
        queryset = queryset.filter(
            Q(PersonalInfo__name__icontains=holder_search) |
            Q(PersonalInfo__family__icontains=holder_search) |
            Q(PersonalInfo__Personnel_number__icontains=holder_search)
        )
    
    # فیلتر بر اساس نوع کالا
    if type_filter:
        queryset = queryset.filter(type_Item=type_filter)
    
    # فیلتر بر اساس وضعیت کالا
    if status_filter:
        queryset = queryset.filter(status_item=status_filter)
    
    # فیلتر بر اساس زیر وضعیت کالا
    if sub_status_filter:
        queryset = queryset.filter(status_sub_item=sub_status_filter)
    
    queryset = queryset.order_by('-register_date')
    
    # ایجاد workbook و worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "لیست ��الاها"
    
    # تنظیم RTL (راست به چپ) برای worksheet
    ws.sheet_view.rightToLeft = True
    
    # تنظیم عرض ستون‌ها
    column_widths = [8, 25, 15, 15, 30, 15, 15, 20, 15, 25, 20, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # استایل هدر
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # هدرهای جدول
    headers = [
        'ردیف', 'نام کالا', 'نوع کالا', 'برند', 'پیکربندی', 
        'وضعیت کالا', 'زیر وضعیت', 'شماره سریال', 'کد محصول', 'دارنده حساب', 
        'تاریخ ثبت', 'تاریخ بروزرسانی'
    ]
    
    # نوشتن هدرها
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # نوشتن داده‌ها
    for row_num, item in enumerate(queryset, 2):
        # ردیف
        ws.cell(row=row_num, column=1, value=row_num - 1)
        
        # نام کالا
        ws.cell(row=row_num, column=2, value=item.Technical_items or 'تعریف نشده')
        
        # نوع کالا
        type_display = 'فنی' if item.type_Item == 'Technical' else 'غیر فنی' if item.type_Item == 'Non-technical' else 'نامشخص'
        ws.cell(row=row_num, column=3, value=type_display)
        
        # برند
        ws.cell(row=row_num, column=4, value=item.brand or 'تعریف نشده')
        
        # پیکربندی
        ws.cell(row=row_num, column=5, value=item.Configuration or 'ندارد')
        
        # وضعیت کالا
        status_display = {
            'hardware': 'سخت افزار (تعمیری)',
            'Delivery': 'تحویل',
            'warehouse': 'انبار',
            'Healthy': 'سالم',
            'Repairing': 'تعمیری', 
            'worn out': 'فرسوده',
            'other': 'سایر'
        }.get(item.status_item, 'نامشخص')
        ws.cell(row=row_num, column=6, value=status_display)
        
        # زیر وضعیت کالا
        sub_status_display = item.get_status_sub_item_display() if item.status_sub_item else 'ندارد'
        ws.cell(row=row_num, column=7, value=sub_status_display)
        
        # شماره سریال
        ws.cell(row=row_num, column=8, value=item.serial_number or 'ندارد')
        
        # کد محصول
        ws.cell(row=row_num, column=9, value=item.Product_code or 'ندارد')
        
        # دارنده حساب
        if item.PersonalInfo:
            holder_name = f"{item.PersonalInfo.name} {item.PersonalInfo.family} ({item.PersonalInfo.Personnel_number})"
        else:
            holder_name = 'بدون دارنده'
        ws.cell(row=row_num, column=10, value=holder_name)
        
        # تاریخ ثبت
        ws.cell(row=row_num, column=11, value=str(item.jinfo))
        
        # تاریخ بروزرسانی
        if item.update_date:
            update_date_str = item.update_date.strftime('%Y/%m/%d %H:%M')
        else:
            update_date_str = 'ندارد'
        ws.cell(row=row_num, column=12, value=update_date_str)
    
    # تنظیم alignment برای همه سلول‌ها
    for row in ws.iter_rows():
        for cell in row:
            if cell.row > 1:  # به جز هدر
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # تنظیم نام فایل
    current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename_parts = ['کالاها']
    
    if search_query:
        filename_parts.append(f'نام_{search_query}')
    if brand_search:
        filename_parts.append(f'برند_{brand_search}')
    if serial_search:
        filename_parts.append(f'سریال_{serial_search}')
    if code_search:
        filename_parts.append(f'کد_{code_search}')
    if holder_search:
        filename_parts.append(f'دارنده_{holder_search}')
    if type_filter:
        type_name = 'فنی' if type_filter == 'Technical' else 'غیرفنی'
        filename_parts.append(f'نوع_{type_name}')
    if status_filter:
        status_names = {
            'hardware': 'سخت_افزار',
            'Delivery': 'تحویل',
            'warehouse': 'انبار'
        }
        filename_parts.append(f'وضعیت_{status_names.get(status_filter, status_filter)}')
    
    if sub_status_filter:
        sub_status_names = {
            'repair': 'تعمیر',
            'upgrade': 'ارتقا',
            'external': 'خارج',
            'internal': 'داخل',
            'ready': 'آماده بکار',
            'returned_good': 'عودتی_سالم',
            'returned_worn': 'عودتی_فرسوده'
        }
        filename_parts.append(f'زیروضعیت_{sub_status_names.get(sub_status_filter, sub_status_filter)}')
    
    filename_parts.append(current_time)
    filename = '_'.join(filename_parts) + '.xlsx'
    
    # ایجاد response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # ذخیره workbook در response
    wb.save(response)
    
    return response


@login_required
def import_excel(request):
    """صفحه آپلود فایل Excel"""
    return render(request, 'registration/import_excel.html')

@login_required
def process_excel(request):
    """پردازش فایل Excel و نمایش پیش‌نمایش برای تأیید"""
    if request.method != 'POST':
        messages.error(request, 'روش درخواست نامعتبر است.')
        return redirect('account:import_excel')
    
    if 'excel_file' not in request.FILES:
        messages.error(request, 'لطفاً فایل Excel را انتخاب کنید.')
        return redirect('account:import_excel')
    
    excel_file = request.FILES['excel_file']
    
    # بررسی فرمت فایل
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'فرمت فایل باید Excel باشد (.xlsx یا .xls)')
        return redirect('account:import_excel')
    
    try:
        # خواندن فایل Excel
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # لیست برای ذخیره داده‌های پردازش شده
        processed_items = []
        errors = []
        
        # خواندن داده‌ها از ردیف دوم (ردیف اول هدر است)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # اگر ردیف خالی باشد
                continue
                
            try:
                # استخراج داده‌ها از ستون‌ها
                item_name = row[1] if len(row) > 1 else None
                item_type = row[2] if len(row) > 2 else None
                brand = row[3] if len(row) > 3 else None
                configuration = row[4] if len(row) > 4 else None
                status = row[5] if len(row) > 5 else None
                serial_number = row[6] if len(row) > 6 else None
                product_code = row[7] if len(row) > 7 else None
                holder_info = row[8] if len(row) > 8 else None
                
                # تبدیل نوع کالا از فارسی به انگلیسی
                type_mapping = {
                    'فنی': 'Technical',
                    'غیر فنی': 'Non-technical'
                }
                item_type_en = type_mapping.get(item_type, 'Technical')
                
                # تبدیل وضعیت از فارسی به انگلیسی
                status_mapping = {
                    'سالم': 'Healthy',
                    'تعمیری': 'Repairing',
                    'فرسوده': 'worn out',
                    'انبار': 'warehouse',
                    'سایر': 'other'
                }
                status_en = status_mapping.get(status, 'Healthy')
                
                # بررسی وجود کالا بر اساس شماره سریال یا کد محصول
                existing_item = None
                action_type = 'create'
                
                if serial_number:
                    existing_item = Items.objects.filter(serial_number=serial_number).first()
                elif product_code:
                    existing_item = Items.objects.filter(Product_code=product_code).first()
                
                if existing_item:
                    action_type = 'update'
                
                # پردازش اطلاعات دارنده
                personal_info = None
                if holder_info and holder_info != 'بدون دارنده':
                    # استخراج شماره پرسنلی از داخل پرانتز
                    import re
                    personnel_match = re.search(r'\((\d+)\)', str(holder_info))
                    if personnel_match:
                        personnel_number = personnel_match.group(1)
                        personal_info = PersonalInfo.objects.filter(Personnel_number=personnel_number).first()
                
                processed_item = {
                    'row_number': row_num,
                    'action_type': action_type,
                    'existing_item_id': existing_item.id if existing_item else None,
                    'data': {
                        'Technical_items': item_name,
                        'type_Item': item_type_en,
                        'brand': brand,
                        'Configuration': configuration,
                        'status_item': status_en,
                        'serial_number': serial_number,
                        'Product_code': product_code,
                        'PersonalInfo_id': personal_info.id if personal_info else None,
                    },
                    'display_data': {
                        'item_name': item_name or 'تعریف نشده',
                        'item_type': item_type or 'فنی',
                        'brand': brand or 'تعریف نشده',
                        'configuration': configuration or 'ندارد',
                        'status': status or 'سالم',
                        'serial_number': serial_number or 'ندارد',
                        'product_code': product_code or 'ندارد',
                        'holder_info': holder_info or 'بدون دارنده',
                    }
                }
                
                processed_items.append(processed_item)
                
            except Exception as e:
                errors.append(f'خطا در ردیف {row_num}: {str(e)}')
        
        if errors:
            for error in errors:
                messages.error(request, error)
        
        if not processed_items:
            messages.warning(request, 'هیچ داده معتبری در فایل یافت نشد.')
            return redirect('account:import_excel')
        
        # ذخیره داده‌ها در session برای مرحله تأیید
        request.session['import_data'] = processed_items
        
        context = {
            'processed_items': processed_items,
            'total_items': len(processed_items),
            'create_count': len([item for item in processed_items if item['action_type'] == 'create']),
            'update_count': len([item for item in processed_items if item['action_type'] == 'update']),
        }
        
        return render(request, 'registration/confirm_import.html', context)
        
    except Exception as e:
        messages.error(request, f'خطا در پردازش فایل: {str(e)}')
        return redirect('account:import_excel')

@login_required
def confirm_import(request):
    """تأیید و اعمال تغییرات"""
    if request.method != 'POST':
        messages.error(request, 'روش درخواست نامعتبر است.')
        return redirect('account:home')
    
    import_data = request.session.get('import_data')
    if not import_data:
        messages.error(request, 'داده‌های وارداتی یافت نشد.')
        return redirect('account:import_excel')
    
    confirmed_items = request.POST.getlist('confirmed_items')
    
    success_count = 0
    error_count = 0
    
    for item_data in import_data:
        row_number = str(item_data['row_number'])
        
        if row_number in confirmed_items:
            try:
                if item_data['action_type'] == 'create':
                    # آماده‌سازی داده‌ها برای ایجاد
                    create_data = item_data['data'].copy()
                    
                    # تبدیل PersonalInfo_id به PersonalInfo object
                    if create_data['PersonalInfo_id']:
                        try:
                            personal_info = PersonalInfo.objects.get(id=create_data['PersonalInfo_id'])
                            create_data['PersonalInfo'] = personal_info
                        except PersonalInfo.DoesNotExist:
                            create_data['PersonalInfo'] = None
                    else:
                        create_data['PersonalInfo'] = None
                    
                    # حذف PersonalInfo_id از داده‌ها
                    del create_data['PersonalInfo_id']
                    
                    # ایجاد کالای جدید
                    Items.objects.create(**create_data)
                    success_count += 1
                    
                elif item_data['action_type'] == 'update':
                    # دریافت کالای موجود
                    existing_item = Items.objects.get(id=item_data['existing_item_id'])
                    
                    # به‌روزرسانی فیلدها
                    for key, value in item_data['data'].items():
                        if key == 'PersonalInfo_id':
                            if value:
                                try:
                                    personal_info = PersonalInfo.objects.get(id=value)
                                    existing_item.PersonalInfo = personal_info
                                except PersonalInfo.DoesNotExist:
                                    existing_item.PersonalInfo = None
                            else:
                                existing_item.PersonalInfo = None
                        elif value is not None:  # فقط مقادیر غیر خالی را به‌روزرسانی کن
                            setattr(existing_item, key, value)
                    
                    existing_item.save()
                    success_count += 1
                    
            except Exception as e:
                error_count += 1
                messages.error(request, f'خطا در پردازش ردیف {row_number}: {str(e)}')
    
    # پاک کردن داده‌ها از session
    if 'import_data' in request.session:
        del request.session['import_data']
    
    if success_count > 0:
        messages.success(request, f'{success_count} مورد با موفقیت ��ردازش شد.')
    
    if error_count > 0:
        messages.warning(request, f'{error_count} مورد با خطا مواجه شد.')
    
    return redirect('account:home')


# ==================== DOCUMENTS VIEWS ====================

@login_required
def documents_list(request):
    """نمایش لیست مدارک"""
    from holder.models import Documents
    
    documents = Documents.objects.all().order_by('-register_date')
    
    # آمار مدارک
    total_documents = documents.count()
    online_documents = documents.filter(Type_of_training='online').count()
    offline_documents = documents.filter(Type_of_training='offline').count()
    
    context = {
        'object_list': documents,
        'total_documents': total_documents,
        'online_documents': online_documents,
        'offline_documents': offline_documents,
    }
    
    return render(request, 'registration/documents_list.html', context)

@login_required
def document_add(request):
    """افزودن مدرک جدید"""
    from holder.forms import DocumentForm
    
    if request.method == 'POST':
        form = DocumentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'مدرک با موفقیت اضافه شد.')
            return redirect('account:documents_list')
        else:
            messages.error(request, 'لطفاً خطاه��ی فرم را بررسی کنید.')
    else:
        form = DocumentForm()
    
    context = {
        'form': form,
        'title': 'افزودن مدرک جدید'
    }
    
    return render(request, 'registration/document_form.html', context)

@login_required
def document_detail(request, pk):
    """جزئیات مدرک"""
    from holder.models import Documents
    document = get_object_or_404(Documents, pk=pk)
    
    context = {
        'document': document,
    }
    
    return render(request, 'registration/document_detail.html', context)

@login_required
def document_edit(request, pk):
    """ویرایش مدرک"""
    from holder.models import Documents
    from holder.forms import DocumentForm
    
    document = get_object_or_404(Documents, pk=pk)
    
    if request.method == 'POST':
        form = DocumentForm(request.POST, instance=document)
        if form.is_valid():
            form.save()
            messages.success(request, 'مدرک با موفقیت به‌روزرسانی شد.')
            return redirect('account:documents_list')
        else:
            messages.error(request, 'لطفاً خطاهای فرم را بررسی کنید.')
    else:
        form = DocumentForm(instance=document)
    
    context = {
        'form': form,
        'document': document,
        'title': 'ویرایش مدرک'
    }
    
    return render(request, 'registration/document_form.html', context)

class DocumentDeleteView(LoginRequiredMixin, DeleteView):
    """حذف مدرک"""
    from holder.models import Documents
    model = Documents
    template_name = 'registration/document_confirm_delete.html'
    success_url = reverse_lazy('account:documents_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'مدرک با موفقیت حذف شد.')
        return super().delete(request, *args, **kwargs)

@login_required
def document_delete(request, pk):
    """حذف مدرک - wrapper function"""
    view = DocumentDeleteView.as_view()
    return view(request, pk=pk)


# ==================== MISSIONS VIEWS ====================

@login_required
def missions_list(request):
    """نمایش لیست ماموریت‌ها"""
    from holder.models import Mission
    
    missions = Mission.objects.all().order_by('-register_date')
    
    # آمار ماموریت‌ها
    total_missions = missions.count()
    active_missions = missions.count()  # فعلاً همه ماموریت‌ها فعال در نظر گرفته می‌شوند
    total_participants = sum([mission.PersonalInfo.count() for mission in missions])
    
    context = {
        'object_list': missions,
        'total_missions': total_missions,
        'active_missions': active_missions,
        'total_participants': total_participants,
    }
    
    return render(request, 'registration/missions_list.html', context)

@login_required
def mission_add(request):
    """افزودن ماموریت جدید"""
    from holder.forms import MissionForm
    
    if request.method == 'POST':
        form = MissionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'ماموریت با موفقیت اضافه شد.')
            return redirect('account:missions_list')
        else:
            messages.error(request, 'لطفاً خطاهای فرم را بررسی کنید.')
    else:
        form = MissionForm()
    
    context = {
        'form': form,
        'title': 'افزودن ماموریت جدید'
    }
    
    return render(request, 'registration/mission_form.html', context)

@login_required
def mission_edit(request, pk):
    """ویرایش ماموریت"""
    from holder.models import Mission
    from holder.forms import MissionForm
    
    mission = get_object_or_404(Mission, pk=pk)
    
    if request.method == 'POST':
        form = MissionForm(request.POST, instance=mission)
        if form.is_valid():
            form.save()
            messages.success(request, 'ماموریت با موفقیت به‌روزرسانی شد.')
            return redirect('account:mission_detail', pk=mission.pk)
        else:
            messages.error(request, 'لطفاً خطاهای فرم را بررسی کنید.')
    else:
        form = MissionForm(instance=mission)
    
    context = {
        'form': form,
        'mission': mission,
        'title': 'ویرایش ماموریت'
    }
    
    return render(request, 'registration/mission_form.html', context)

@login_required
def mission_delete(request, pk):
    """حذف ماموریت"""
    from holder.models import Mission
    
    mission = get_object_or_404(Mission, pk=pk)
    
    if request.method == 'POST':
        mission.delete()
        messages.success(request, 'ماموریت با موفقیت حذف شد.')
        return redirect('account:missions_list')
    
    context = {
        'object': mission,
        'title': 'حذف ماموریت'
    }
    
    return render(request, 'registration/mission_confirm_delete.html', context)


# ==================== RESULTS VIEWS ====================

@login_required
def results_list(request):
    """نمایش لیست نتایج"""
    from holder.models import Results
    from django.utils import timezone
    
    results = Results.objects.select_related('PersonalInfo').all().order_by('-register_date')
    
    # آمار نتایج
    total_results = results.count()
    total_meetings = sum([result.Internal_meetings for result in results])
    recent_results = results.filter(register_date__gte=timezone.now() - timezone.timedelta(days=30)).count()
    
    context = {
        'object_list': results,
        'total_results': total_results,
        'total_meetings': total_meetings,
        'recent_results': recent_results,
    }
    
    return render(request, 'registration/results_list.html', context)

@login_required
def result_add(request):
    """افزودن نتیجه جدید"""
    from holder.forms import ResultForm
    
    if request.method == 'POST':
        form = ResultForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'نتیجه با موفقیت اضافه شد.')
            return redirect('account:results_list')
        else:
            messages.error(request, 'لطفاً خطاهای فرم را بررسی کنید.')
    else:
        form = ResultForm()
    
    context = {
        'form': form,
        'title': 'افزودن نتیجه جدید'
    }
    
    return render(request, 'registration/result_form.html', context)

@login_required
def result_edit(request, pk):
    """ویرایش نتیجه"""
    from holder.models import Results
    from holder.forms import ResultForm
    
    result = get_object_or_404(Results, pk=pk)
    
    if request.method == 'POST':
        form = ResultForm(request.POST, instance=result)
        if form.is_valid():
            form.save()
            messages.success(request, 'نتیجه با موفقیت به‌روزرسانی شد.')
            return redirect('account:results_list')
        else:
            messages.error(request, 'لطفاً خطاهارا فرم را بررسی کنید.')
    else:
        form = ResultForm(instance=result)
    
    context = {
        'form': form,
        'result': result,
        'title': 'ویرایش نتیجه'
    }
    
    return render(request, 'registration/result_form.html', context)

class ResultDeleteView(LoginRequiredMixin, DeleteView):
    """حذف نتیجه"""
    from holder.models import Results
    model = Results
    template_name = 'registration/result_confirm_delete.html'
    success_url = reverse_lazy('account:results_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'نتیجه با موفقیت حذف شد.')
        return super().delete(request, *args, **kwargs)

@login_required
def result_delete(request, pk):
    """حذف نتیجه - wrapper function"""
    view = ResultDeleteView.as_view()
    return view(request, pk=pk)