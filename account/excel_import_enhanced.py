from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from holder.models import Items, PersonalInfo
import openpyxl
import json

@login_required
def process_excel_enhanced(request):
    """پردازش فایل Excel با تأیید فیلد به فیلد"""
    if request.method != 'POST':
        messages.error(request, 'روش درخواست نامعتبر است.')
        return redirect('account:import_excel')
    
    if 'excel_file' not in request.FILES:
        messages.error(request, 'لطفاً فایل Excel را انتخاب کنید.')
        return redirect('account:import_excel')
    
    excel_file = request.FILES['excel_file']
    
    # بررسی فرمت فایل
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'فرمت فایل باید Excel باش�� (.xlsx یا .xls)')
        return redirect('account:import_excel')
    
    try:
        # خواندن فایل Excel
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # لیست برای ذخیره داده‌های پردازش شده
        processed_items = []
        errors = []
        
        # خواندن هدرها از ردیف اول
        headers = []
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row:
            headers = [str(cell).strip() if cell else '' for cell in header_row]
        
        # تعریف نقشه‌برداری هدرها به فیلدهای مدل
        header_mapping = {
            # نام کالا
            'نام کالا': 'item_name',
            'نام': 'item_name',
            'کالا': 'item_name',
            'Technical_items': 'item_name',
            
            # نوع کالا
            'نوع کالا': 'item_type',
            'نوع': 'item_type',
            'type_Item': 'item_type',
            
            # برند
            'برند': 'brand',
            'brand': 'brand',
            
            # پیکربندی
            'پیکربندی': 'configuration',
            'Configuration': 'configuration',
            'تنظیمات': 'configuration',
            
            # وضعیت اصلی
            'وضعیت کالا': 'status_main',
            'وضعیت اصلی': 'status_main',
            'وضعیت': 'status_main',
            'status_item': 'status_main',
            
            # زیر وضعیت
            'زیر وضعیت': 'status_sub',
            'زیرمجموعه وضعیت': 'status_sub',
            'status_sub_item': 'status_sub',
            
            # شماره سریال
            'شماره سریال': 'serial_number',
            'سریال': 'serial_number',
            'serial_number': 'serial_number',
            
            # کد محصول
            'کد محصول': 'product_code',
            'کد کالا': 'product_code',
            'کد': 'product_code',
            'Product_code': 'product_code',
            
            # دارنده
            'دارنده': 'holder_info',
            'دارنده حساب': 'holder_info',
            'PersonalInfo': 'holder_info',
            'شخص': 'holder_info',
            
            # تعداد
            'تعداد': 'number',
            'تعداد کالا': 'number',
            'Number': 'number'
        }
        
        # ایجاد نقشه ایندکس ستون‌ها
        column_indices = {}
        for i, header in enumerate(headers):
            if header in header_mapping:
                column_indices[header_mapping[header]] = i
        
        # خواندن داده‌ها از ردیف دوم (ردیف اول هدر است)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # اگر ردیف خالی باشد
                continue
                
            try:
                # استخراج داده‌ها بر اساس نقشه ستون‌ها
                def get_cell_value(field_name):
                    """دریافت مقدار سلول بر اساس نام فیلد"""
                    if field_name in column_indices:
                        index = column_indices[field_name]
                        if index < len(row):
                            value = row[index]
                            # تبدیل مقادیر خالی به None
                            if value is None or str(value).strip() == '':
                                return None
                            return str(value).strip()
                    return None
                
                item_name = get_cell_value('item_name')
                item_type = get_cell_value('item_type')
                brand = get_cell_value('brand')
                configuration = get_cell_value('configuration')
                status_main = get_cell_value('status_main')
                status_sub = get_cell_value('status_sub')
                serial_number = get_cell_value('serial_number')
                product_code = get_cell_value('product_code')
                holder_info = get_cell_value('holder_info')
                number = get_cell_value('number')
                
                # تبدیل نوع کالا از فارسی به انگلیسی
                type_mapping = {
                    'فنی': 'Technical',
                    'غیر فنی': 'Non-technical'
                }
                item_type_en = type_mapping.get(item_type, 'Technical')
                
                # تبدیل وضعیت اصلی از فارسی به انگلیسی
                status_main_mapping = {
                    'سخت افزار (تعمیری)': 'hardware',
                    'تعمیری': 'hardware',
                    'سخت افزار': 'hardware',
                    'تحویل': 'Delivery',
                    'انبار': 'warehouse'
                }
                status_main_en = status_main_mapping.get(status_main, 'warehouse')
                
                # تبدیل زیر وضعیت از فارسی به انگلیسی
                status_sub_mapping = {
                    'تعمیر': 'repair',
                    'ارتقا': 'upgrade',
                    'خارج': 'external',
                    'داخل': 'internal',
                    'آماده بکار': 'ready',
                    'عودتی سالم': 'returned_good',
                    'عودتی فرسوده': 'returned_worn'
                }
                status_sub_en = status_sub_mapping.get(status_sub, None) if status_sub else None
                
                # بررسی وجود کالا بر اساس شماره سریال یا کد محصول
                existing_item = None
                action_type = 'create'
                
                if serial_number:
                    existing_item = Items.objects.filter(serial_number=serial_number).first()
                elif product_code:
                    existing_item = Items.objects.filter(Product_code=product_code).first()
                
                if existing_item:
                    action_type = 'update'
                
                # پردازش اطلاعات دارنده
                personal_info = None
                if holder_info and holder_info != 'بدون دارنده':
                    # استخراج شماره پرسنلی از داخل پرانتز
                    import re
                    personnel_match = re.search(r'\((\d+)\)', str(holder_info))
                    if personnel_match:
                        personnel_number = personnel_match.group(1)
                        personal_info = PersonalInfo.objects.filter(Personnel_number=personnel_number).first()
                
                # ایجاد ساختار داده با فیلدهای قابل تأیید جداگانه
                field_confirmations = {
                    'Technical_items': {
                        'value': item_name,
                        'display': item_name or 'تعریف نشده',
                        'confirmed': bool(item_name),
                        'required': True,
                        'label': 'نام کالا',
                        'type': 'text'
                    },
                    'type_Item': {
                        'value': item_type_en,
                        'display': item_type or 'فنی',
                        'confirmed': bool(item_type),
                        'required': True,
                        'label': 'نوع کالا',
                        'type': 'select',
                        'options': [
                            {'value': 'Technical', 'label': 'فنی'},
                            {'value': 'Non-technical', 'label': 'غیر فنی'}
                        ]
                    },
                    'brand': {
                        'value': brand,
                        'display': brand or 'تعریف نشده',
                        'confirmed': bool(brand),
                        'required': False,
                        'label': 'برند',
                        'type': 'text'
                    },
                    'Configuration': {
                        'value': configuration,
                        'display': configuration or 'ندارد',
                        'confirmed': bool(configuration),
                        'required': False,
                        'label': 'پیکربندی',
                        'type': 'textarea'
                    },
                    'status_item': {
                        'value': status_main_en,
                        'display': status_main or 'انبار',
                        'confirmed': bool(status_main),
                        'required': True,
                        'label': 'وضعیت اصلی',
                        'type': 'select',
                        'options': [
                            {'value': 'hardware', 'label': 'سخت افزار (تعمیری)'},
                            {'value': 'Delivery', 'label': 'تحویل'},
                            {'value': 'warehouse', 'label': 'انبار'}
                        ]
                    },
                    'status_sub_item': {
                        'value': status_sub_en,
                        'display': status_sub or 'ندارد',
                        'confirmed': bool(status_sub_en),
                        'required': False,
                        'label': 'زیر وضعیت',
                        'type': 'select',
                        'options': []  # Will be populated dynamically based on main status
                    },
                    'serial_number': {
                        'value': serial_number,
                        'display': serial_number or 'ندارد',
                        'confirmed': bool(serial_number),
                        'required': False,
                        'label': 'شماره سریال',
                        'type': 'text'
                    },
                    'Product_code': {
                        'value': product_code,
                        'display': product_code or 'ندارد',
                        'confirmed': bool(product_code),
                        'required': False,
                        'label': 'کد محصول',
                        'type': 'text'
                    },
                    'PersonalInfo': {
                        'value': personal_info.Personnel_number if personal_info else None,
                        'display': holder_info or 'بدون دارنده',
                        'confirmed': bool(personal_info),
                        'required': False,
                        'label': 'دارنده حساب',
                        'type': 'select',
                        'options': []  # Will be populated with all PersonalInfo
                    },
                    'Number': {
                        'value': int(number) if number and str(number).isdigit() else 1,
                        'display': number or '1',
                        'confirmed': bool(number),
                        'required': False,
                        'label': 'تعداد کالا',
                        'type': 'number'
                    }
                }
                
                # پر کردن گزینه‌های زیر وضعیت بر اساس وضعیت اصلی
                if status_main_en == 'hardware':
                    field_confirmations['status_sub_item']['options'] = [
                        {'value': 'repair', 'label': 'تعمیر'},
                        {'value': 'upgrade', 'label': 'ارتقا'}
                    ]
                elif status_main_en == 'Delivery':
                    field_confirmations['status_sub_item']['options'] = [
                        {'value': 'external', 'label': 'خارج'},
                        {'value': 'internal', 'label': 'داخل'}
                    ]
                elif status_main_en == 'warehouse':
                    field_confirmations['status_sub_item']['options'] = [
                        {'value': 'ready', 'label': 'آماده بکار'},
                        {'value': 'returned_good', 'label': 'عودتی سالم'},
                        {'value': 'returned_worn', 'label': 'عودتی فرسوده'}
                    ]
                
                processed_item = {
                    'row_number': row_num,
                    'action_type': action_type,
                    'existing_item_id': existing_item.id if existing_item else None,
                    'existing_item_name': existing_item.Technical_items if existing_item else None,
                    'field_confirmations': field_confirmations,
                    'has_warnings': any(not field['confirmed'] and field['value'] for field in field_confirmations.values()),
                    'has_errors': any(not field['value'] and field['required'] for field in field_confirmations.values())
                }
                
                processed_items.append(processed_item)
                
            except Exception as e:
                errors.append(f'خطا در ردیف {row_num}: {str(e)}')
        
        if errors:
            for error in errors:
                messages.error(request, error)
        
        if not processed_items:
            messages.warning(request, 'هیچ داده معتبری در فایل یافت نشد.')
            return redirect('account:import_excel')
        
        # دریافت تمام افراد برای انتخاب دارنده
        all_people = PersonalInfo.objects.all().order_by('name', 'family')
        people_options = [{'value': '', 'label': 'بدون دارنده'}]
        for person in all_people:
            people_options.append({
                'value': person.Personnel_number,
                'label': f"{person.name} {person.family} ({person.Personnel_number})"
            })
        
        # اضافه کردن گزینه‌های افراد به همه آیتم‌ها
        for item in processed_items:
            item['field_confirmations']['PersonalInfo']['options'] = people_options
        
        # ذخیره داده‌ها در session برای مرحله تأیید
        request.session['import_data'] = processed_items
        
        context = {
            'processed_items': processed_items,
            'total_items': len(processed_items),
            'create_count': len([item for item in processed_items if item['action_type'] == 'create']),
            'update_count': len([item for item in processed_items if item['action_type'] == 'update']),
            'warning_count': len([item for item in processed_items if item['has_warnings']]),
            'error_count': len([item for item in processed_items if item['has_errors']])
        }
        
        return render(request, 'registration/confirm_import_enhanced.html', context)
        
    except Exception as e:
        messages.error(request, f'خطا در پردازش فایل: {str(e)}')
        return redirect('account:import_excel')

@login_required
def confirm_import_enhanced(request):
    """تأیید و اعمال تغییرات با تأیید فیلد به فیلد"""
    if request.method != 'POST':
        messages.error(request, 'روش درخواست نامعتبر است.')
        return redirect('account:home')
    
    import_data = request.session.get('import_data')
    if not import_data:
        messages.error(request, 'داده‌های وارداتی یافت نشد.')
        return redirect('account:import_excel')
    
    # دریافت داده‌های فرم
    confirmed_items = request.POST.getlist('confirmed_items')
    
    success_count = 0
    error_count = 0
    
    for item_data in import_data:
        row_number = str(item_data['row_number'])
        
        if row_number in confirmed_items:
            try:
                # جمع‌آوری داده‌های تأیید شده از فرم
                confirmed_data = {}
                
                for field_name, field_info in item_data['field_confirmations'].items():
                    form_field_name = f"field_{row_number}_{field_name}"
                    
                    if form_field_name in request.POST:
                        value = request.POST[form_field_name]
                        
                        # پردازش مقادیر خاص
                        if field_name == 'PersonalInfo':
                            if value:
                                try:
                                    personal_info = PersonalInfo.objects.get(Personnel_number=value)
                                    confirmed_data['PersonalInfo'] = personal_info
                                except PersonalInfo.DoesNotExist:
                                    confirmed_data['PersonalInfo'] = None
                            else:
                                confirmed_data['PersonalInfo'] = None
                        elif value:  # فقط مقادیر غیر خالی
                            confirmed_data[field_name] = value
                
                if item_data['action_type'] == 'create':
                    # ایجاد کالای جدید
                    Items.objects.create(**confirmed_data)
                    success_count += 1
                    
                elif item_data['action_type'] == 'update':
                    # دریافت کالای موجود و به‌روزرسانی
                    existing_item = Items.objects.get(id=item_data['existing_item_id'])
                    
                    for key, value in confirmed_data.items():
                        setattr(existing_item, key, value)
                    
                    existing_item.save()
                    success_count += 1
                    
            except Exception as e:
                error_count += 1
                messages.error(request, f'خطا در پردازش ردیف {row_number}: {str(e)}')
    
    # پاک کردن داده‌ها از session
    if 'import_data' in request.session:
        del request.session['import_data']
    
    if success_count > 0:
        messages.success(request, f'{success_count} مورد با موفقیت پردازش شد.')
    
    if error_count > 0:
        messages.warning(request, f'{error_count} مورد با خطا مواجه شد.')
    
    return redirect('account:home')

@login_required
def get_sub_status_options(request):
    """دریافت گزینه‌های زیر وضعیت بر اساس وضعیت اصلی"""
    main_status = request.GET.get('main_status')
    
    options = []
    if main_status == 'hardware':
        options = [
            {'value': 'repair', 'label': 'تعمیر'},
            {'value': 'upgrade', 'label': 'ارتقا'}
        ]
    elif main_status == 'Delivery':
        options = [
            {'value': 'external', 'label': 'خارج'},
            {'value': 'internal', 'label': 'داخل'}
        ]
    elif main_status == 'warehouse':
        options = [
            {'value': 'ready', 'label': 'آماده بکار'},
            {'value': 'returned_good', 'label': 'عودتی سالم'},
            {'value': 'returned_worn', 'label': 'عودتی فرسوده'}
        ]
    
    return JsonResponse({'options': options})