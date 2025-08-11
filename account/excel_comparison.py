from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from holder.models import Items, PersonalInfo
import openpyxl
import json

@login_required
def compare_excel_with_items(request):
    """مقایسه فایل Excel با کالاهای موجود در سیستم"""
    if request.method != 'POST':
        messages.error(request, 'روش درخواست نامعتبر است.')
        return redirect('account:import_excel')
    
    if 'excel_file' not in request.FILES:
        messages.error(request, 'لطفاً فایل Excel را انتخاب کنید.')
        return redirect('account:import_excel')
    
    excel_file = request.FILES['excel_file']
    comparison_type = request.POST.get('comparison_type', 'all')
    
    # بررسی فرمت فایل
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'فرمت فایل باید Excel باشد (.xlsx یا .xls)')
        return redirect('account:import_excel')
    
    try:
        # خواندن فایل Excel
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # لیست برای ذخیره نتایج مقایسه
        comparison_results = {
            'new_items': [],           # کالاهای جدید (در Excel هست، در سیستم نیست)
            'existing_items': [],      # کالاهای موجود (در هر دو هست)
            'differences': [],         # کالاهایی که تفاوت دارند
            'system_only': [],         # کالاهایی که فقط در سیستم هستند
            'excel_errors': []         # خطاهای فایل Excel
        }
        
        # خواندن هدرها از ردیف اول
        headers = []
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row:
            headers = [str(cell).strip() if cell else '' for cell in header_row]
        
        # تعریف نقشه‌برداری هدرها به فیلدهای مدل
        header_mapping = {
            # نام کالا
            'نام کالا': 'item_name',
            'نام': 'item_name',
            'کالا': 'item_name',
            'Technical_items': 'item_name',
            
            # نوع کالا
            'نوع کالا': 'item_type',
            'نوع': 'item_type',
            'type_Item': 'item_type',
            
            # برند
            'برند': 'brand',
            'brand': 'brand',
            
            # پیکربندی
            'پیکربندی': 'configuration',
            'Configuration': 'configuration',
            'تنظیمات': 'configuration',
            
            # وضعیت اصلی
            'وضعیت کالا': 'status_main',
            'وضعیت اصلی': 'status_main',
            'وضعیت': 'status_main',
            'status_item': 'status_main',
            
            # زیر وضعیت
            'زیر وضعیت': 'status_sub',
            'زیرمجموعه وضعیت': 'status_sub',
            'status_sub_item': 'status_sub',
            
            # شماره سریال
            'شماره سریال': 'serial_number',
            'سریال': 'serial_number',
            'serial_number': 'serial_number',
            
            # کد محصول
            'کد محصول': 'product_code',
            'کد کالا': 'product_code',
            'کد': 'product_code',
            'Product_code': 'product_code',
            
            # دارنده
            'دارنده': 'holder_info',
            'دارنده حساب': 'holder_info',
            'PersonalInfo': 'holder_info',
            'شخص': 'holder_info',
            
            # تعداد
            'تعداد': 'number',
            'تعداد کالا': 'number',
            'Number': 'number'
        }
        
        # ایجاد نقشه ایندکس ستون‌ها
        column_indices = {}
        for i, header in enumerate(headers):
            if header in header_mapping:
                column_indices[header_mapping[header]] = i
        
        # مجموعه‌ای از کدهای محصول و شماره سریال‌های موجود در Excel
        excel_product_codes = set()
        excel_serial_numbers = set()
        
                
        # خواندن داده‌ها از ردیف دوم (ردیف اول هدر است)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # اگر ردیف خالی باشد
                continue
                
            try:
                # استخراج داده‌ها بر اساس نقشه ستون‌ها
                def get_cell_value(field_name):
                    """دریافت مقدار سلول بر اساس نام فیلد"""
                    if field_name in column_indices:
                        index = column_indices[field_name]
                        if index < len(row):
                            value = row[index]
                            if value is None or str(value).strip() == '':
                                return None
                            return str(value).strip()
                    return None
                
                item_name = get_cell_value('item_name')
                item_type = get_cell_value('item_type')
                brand = get_cell_value('brand')
                configuration = get_cell_value('configuration')
                status_main = get_cell_value('status_main')
                status_sub = get_cell_value('status_sub')
                serial_number = get_cell_value('serial_number')
                product_code = get_cell_value('product_code')
                holder_info = get_cell_value('holder_info')
                number = get_cell_value('number')
                
                # بررسی وجود کد محصول (ضروری)
                if not product_code:
                    comparison_results['excel_errors'].append({
                        'row': row_num,
                        'error': 'کد محصول وارد نشده است',
                        'data': {
                            'item_name': item_name,
                            'serial_number': serial_number
                        }
                    })
                    continue
                
                # تشخیص نوع کالا از Excel
                excel_item_type_raw = item_type  # مقدار خام از Excel
                
                # اضافه کردن به مجموعه‌های Excel
                excel_product_codes.add(product_code)
                # فقط برای کالاهای فنی شماره سریال را اضافه کن
                if serial_number and (excel_item_type_raw and excel_item_type_raw.lower() in ['technical', 'فنی']):
                    excel_serial_numbers.add(serial_number)
                
                # جستجوی کالا در سیستم بر اساس نوع کالا
                existing_item = None
                is_technical = False
                
                # تشخیص اینکه کالا فنی است یا غیر فنی
                if excel_item_type_raw:
                    if excel_item_type_raw.lower() in ['technical', 'فنی']:
                        is_technical = True
                    elif excel_item_type_raw.lower() in ['non-technical', 'غیر فنی']:
                        is_technical = False
                else:
                    # اگر نوع مشخص نباشد، فرض می‌کنیم فنی است
                    is_technical = True
                
                # تابع برای تبدیل مقادیر Excel به نمایش فارسی
                def convert_excel_to_display(field_type, value):
                    """تبدیل مقادیر Excel به نمایش فارسی"""
                    if not value:
                        return value
                    
                    if field_type == 'item_type':
                        type_mapping = {
                            'Technical': 'فنی',
                            'Non-technical': 'غیر فنی',
                            'فنی': 'فنی',
                            'غیر فنی': 'غیر فنی'
                        }
                        return type_mapping.get(value, value)
                    
                    elif field_type == 'status_main':
                        status_mapping = {
                            'hardware': 'سخت افزار',
                            'Delivery': 'تحویل',
                            'warehouse': 'انبار',
                            'سخت افزار ': 'سخت افزار',
                            'تحویل': 'تحویل',
                            'انبار': 'انبار'
                        }
                        return status_mapping.get(value, value)
                    
                    elif field_type == 'status_sub':
                        sub_status_mapping = {
                            'repair': 'تعمیر',
                            'upgrade': 'ارتقا',
                            'external': 'خارج',
                            'internal': 'داخل',
                            'ready': 'آماده بکار',
                            'returned_good': 'عودتی سالم',
                            'returned_worn': 'عودتی فرسوده',
                            'تعمیر': 'تعمیر',
                            'ارتقا': 'ارتقا',
                            'خارج': 'خارج',
                            'داخل': 'داخل',
                            'آماده بکار': 'آماده بکار',
                            'عودتی سالم': 'عودتی سالم',
                            'عودتی فرسوده': 'عودتی فرسوده'
                        }
                        return sub_status_mapping.get(value, value)
                    
                    return value

                # ساختار داده Excel با تبدیل به نمایش فارسی
                # برای کالاهای غیر فنی، شماره سریال همیشه None است
                final_serial_number = serial_number if is_technical else None
                
                excel_item_data = {
                    'row': row_num,
                    'item_name': item_name,
                    'item_type': convert_excel_to_display('item_type', item_type),
                    'brand': brand,
                    'configuration': configuration,
                    'status_main': convert_excel_to_display('status_main', status_main),
                    'status_sub': convert_excel_to_display('status_sub', status_sub),
                    'serial_number': final_serial_number,
                    'product_code': product_code,
                    'holder_info': holder_info,
                    'number': int(number) if number and str(number).isdigit() else 1,
                    'is_technical': is_technical
                }
                
                if is_technical:
                    # کالاهای فنی: پردازش فوری
                    # برای کالاهای فنی، هم شماره سریال و هم کد محصول باید منحصر به فرد باشند
                    existing_item = None
                    
                    # ابتدا بر اساس شماره سریال جستجو کن (اگر موجود باشد)
                    if serial_number:
                        existing_item = Items.objects.filter(serial_number=serial_number).first()
                    
                    # اگر بر اساس شماره سریال پیدا نشد، بر اساس کد محصول جستجو کن
                    if not existing_item:
                        existing_item = Items.objects.filter(Product_code=product_code).first()
                    
                    # اگر کالا پیدا شد، باید هم شماره سریال و هم کد محصول مطابقت داشته باشند
                    if existing_item:
                        # بررسی مطابقت کامل شماره سریال و کد محصول
                        serial_match = True
                        product_code_match = True
                        
                        # اگر Excel شماره سریال دارد، باید با سیستم مطابقت داشته باشد
                        if serial_number:
                            serial_match = (existing_item.serial_number == serial_number)
                        
                        # کد محصول باید حتماً مطابقت داشته باشد
                        product_code_match = (existing_item.Product_code == product_code)
                        
                        # اگر هر کدام مطابقت نداشته باشد، کالا جدید محسوب می‌شود
                        if not (serial_match and product_code_match):
                            existing_item = None
                    
                    # پردازش کالای فنی
                    if existing_item:
                        # کالا در سیستم موجود است
                        system_item_data = {
                            'id': existing_item.id,
                            'item_name': existing_item.Technical_items,
                            'item_type': existing_item.get_type_Item_display(),
                            'brand': existing_item.brand,
                            'configuration': existing_item.Configuration,
                            'status_main': existing_item.get_status_item_display(),
                            'status_sub': existing_item.get_status_sub_item_display() if existing_item.status_sub_item else None,
                            'serial_number': existing_item.serial_number,
                            'product_code': existing_item.Product_code,
                            'holder_info': f"{existing_item.PersonalInfo.name} {existing_item.PersonalInfo.family} ({existing_item.PersonalInfo.Personnel_number})" if existing_item.PersonalInfo else None,
                            'number': existing_item.Number
                        }
                        
                        # بررسی تفاوت‌ها برای کالاهای فنی
                        differences = []
                        field_comparisons = [
                            ('item_name', Items._meta.get_field('Technical_items').verbose_name),
                            ('item_type', Items._meta.get_field('type_Item').verbose_name),
                            ('brand', Items._meta.get_field('brand').verbose_name),
                            ('configuration', Items._meta.get_field('Configuration').verbose_name),
                            ('status_main', Items._meta.get_field('status_item').verbose_name),
                            ('status_sub', Items._meta.get_field('status_sub_item').verbose_name),
                            ('serial_number', Items._meta.get_field('serial_number').verbose_name),
                            ('product_code', Items._meta.get_field('Product_code').verbose_name),
                            ('number', Items._meta.get_field('Number').verbose_name)
                        ]
                        
                        for field, label in field_comparisons:
                            excel_value = excel_item_data.get(field)
                            system_value = system_item_data.get(field)
                            
                            if excel_value is None:
                                excel_value = ''
                            if system_value is None:
                                system_value = ''
                                
                            if str(excel_value).strip() != str(system_value).strip():
                                differences.append({
                                    'field': field,
                                    'field_label': label,
                                    'excel_value': excel_value,
                                    'system_value': system_value
                                })
                        
                        if differences:
                            comparison_results['differences'].append({
                                'excel_data': excel_item_data,
                                'system_data': system_item_data,
                                'differences': differences
                            })
                        else:
                            comparison_results['existing_items'].append({
                                'excel_data': excel_item_data,
                                'system_data': system_item_data
                            })
                    else:
                        # کالای فنی جدید
                        excel_item_data['new_item_reason'] = f"کالا جدید است - {'شماره سریال' if serial_number else 'کد محصول'} '{serial_number or product_code}' در سیستم موجود نیست"
                        comparison_results['new_items'].append(excel_item_data)
                        
                else:
                    # کالاهای غیر فنی: پردازش فوری هر کالا به صورت جداگانه
                    # برای کالاهای غیر ��نی، فقط بر اساس کد محصول جستجو می‌کنیم
                    existing_items = Items.objects.filter(
                        Product_code=product_code,
                        type_Item='Non-technical'
                    )
                    
                    if existing_items.exists():
                        # کالا در سیستم موجود است - هر کالا را جداگانه بررسی می‌کنیم
                        # آیتم نماینده از سیستم
                        representative_item = existing_items.first()
                        system_item_data = {
                            'id': representative_item.id,
                            'item_name': representative_item.Technical_items,
                            'item_type': representative_item.get_type_Item_display(),
                            'brand': representative_item.brand,
                            'configuration': representative_item.Configuration,
                            'status_main': representative_item.get_status_item_display(),
                            'status_sub': representative_item.get_status_sub_item_display() if representative_item.status_sub_item else None,
                            'serial_number': representative_item.serial_number,
                            'product_code': representative_item.Product_code,
                            'holder_info': f"{representative_item.PersonalInfo.name} {representative_item.PersonalInfo.family} ({representative_item.PersonalInfo.Personnel_number})" if representative_item.PersonalInfo else None,
                            'number': representative_item.Number
                        }
                        
                        # بررسی تفاوت‌ها برای کالاهای غیر فنی
                        differences = []
                        field_comparisons = [
                            ('item_name', Items._meta.get_field('Technical_items').verbose_name),
                            ('item_type', Items._meta.get_field('type_Item').verbose_name),
                            ('brand', Items._meta.get_field('brand').verbose_name),
                            ('configuration', Items._meta.get_field('Configuration').verbose_name),
                            ('status_main', Items._meta.get_field('status_item').verbose_name),
                            ('status_sub', Items._meta.get_field('status_sub_item').verbose_name),
                            ('product_code', Items._meta.get_field('Product_code').verbose_name),
                            ('number', Items._meta.get_field('Number').verbose_name)
                        ]
                        
                        for field, label in field_comparisons:
                            excel_value = excel_item_data.get(field)
                            system_value = system_item_data.get(field)
                            
                            if excel_value is None:
                                excel_value = ''
                            if system_value is None:
                                system_value = ''
                                
                            if str(excel_value).strip() != str(system_value).strip():
                                differences.append({
                                    'field': field,
                                    'field_label': label,
                                    'excel_value': excel_value,
                                    'system_value': system_value
                                })
                        
                        if differences:
                            comparison_results['differences'].append({
                                'excel_data': excel_item_data,
                                'system_data': system_item_data,
                                'differences': differences
                            })
                        else:
                            comparison_results['existing_items'].append({
                                'excel_data': excel_item_data,
                                'system_data': system_item_data
                            })
                    else:
                        # کالای غیر فنی جدید
                        excel_item_data['new_item_reason'] = f"کالا جدید است - کد محصول '{product_code}' در کالاهای غیر فنی موجود نیست"
                        comparison_results['new_items'].append(excel_item_data)
                    
            except Exception as e:
                comparison_results['excel_errors'].append({
                    'row': row_num,
                    'error': f'خطا در پردازش: {str(e)}',
                    'data': {}
                })
        
                
        # یافتن کالاهایی که فقط در سیستم هستند
        if comparison_type in ['all', 'system_only']:
            all_system_items = Items.objects.all()
            for system_item in all_system_items:
                # بررسی اینکه آیا این کالا در Excel موجود است یا نه
                found_in_excel = False
                
                # برای کالاهای فنی: بررسی بر اساس شماره سریال یا کد محصول
                if system_item.type_Item == 'Technical':
                    if system_item.serial_number and system_item.serial_number in excel_serial_numbers:
                        found_in_excel = True
                    elif system_item.Product_code in excel_product_codes:
                        found_in_excel = True
                
                # برای کالاهای غیر فنی: فقط بررسی بر اساس کد محصول
                elif system_item.type_Item == 'Non-technical':
                    if system_item.Product_code in excel_product_codes:
                        found_in_excel = True
                
                if not found_in_excel:
                    comparison_results['system_only'].append({
                        'id': system_item.id,
                        'item_name': system_item.Technical_items,
                        'item_type': system_item.get_type_Item_display(),  # نمایش فارسی
                        'brand': system_item.brand,
                        'configuration': system_item.Configuration,
                        'status_main': system_item.get_status_item_display(),  # نمایش فارسی
                        'status_sub': system_item.get_status_sub_item_display() if system_item.status_sub_item else None,  # نمایش فارسی
                        'serial_number': system_item.serial_number,
                        'product_code': system_item.Product_code,
                        'holder_info': f"{system_item.PersonalInfo.name} {system_item.PersonalInfo.family} ({system_item.PersonalInfo.Personnel_number})" if system_item.PersonalInfo else None,
                        'number': system_item.Number
                    })
        
        # فیلتر کردن نتایج بر اساس نوع مقایسه
        filtered_results = {}
        if comparison_type == 'all':
            filtered_results = comparison_results
        elif comparison_type == 'new_items':
            filtered_results = {
                'new_items': comparison_results['new_items'],
                'excel_errors': comparison_results['excel_errors']
            }
        elif comparison_type == 'existing_items':
            filtered_results = {
                'existing_items': comparison_results['existing_items'],
                'excel_errors': comparison_results['excel_errors']
            }
        elif comparison_type == 'differences':
            filtered_results = {
                'differences': comparison_results['differences'],
                'excel_errors': comparison_results['excel_errors']
            }
        
        # آمار کلی
        stats = {
            'total_excel_items': len(comparison_results['new_items']) + len(comparison_results['existing_items']) + len(comparison_results['differences']),
            'new_items_count': len(comparison_results['new_items']),
            'existing_items_count': len(comparison_results['existing_items']),
            'differences_count': len(comparison_results['differences']),
            'system_only_count': len(comparison_results['system_only']),
            'excel_errors_count': len(comparison_results['excel_errors']),
            'comparison_type': comparison_type
        }
        
        # اضافه کردن اطلاعات فیلدهای مدل برای استفاده در template
        model_fields = {
            'Technical_items': Items._meta.get_field('Technical_items').verbose_name,
            'type_Item': Items._meta.get_field('type_Item').verbose_name,
            'brand': Items._meta.get_field('brand').verbose_name,
            'Product_code': Items._meta.get_field('Product_code').verbose_name,
            'serial_number': Items._meta.get_field('serial_number').verbose_name,
            'status_item': Items._meta.get_field('status_item').verbose_name,
            'Number': Items._meta.get_field('Number').verbose_name,
            'PersonalInfo': Items._meta.get_field('PersonalInfo').verbose_name,
            'Configuration': Items._meta.get_field('Configuration').verbose_name,
            'status_sub_item': Items._meta.get_field('status_sub_item').verbose_name,
        }
        
        context = {
            'comparison_results': filtered_results,
            'stats': stats,
            'comparison_type': comparison_type,
            'comparison_type_display': {
                'all': 'مقایسه کامل',
                'new_items': 'فقط کالاهای جدید',
                'existing_items': 'فقط کالاهای موجود',
                'differences': 'فقط تفاوت‌ها'
            }.get(comparison_type, 'نامشخص'),
            'model_fields': model_fields
        }
        
        # ذخیره داده‌های مقایسه در session برای استفاده در ویرایش
        request.session['comparison_excel_data'] = comparison_results
        
        return render(request, 'registration/excel_comparison_results.html', context)
        
    except Exception as e:
        messages.error(request, f'خطا در پردازش فایل: {str(e)}')
        return redirect('account:import_excel')